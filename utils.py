#!C:\Python27
# -*- coding: utf-8 -*-
import chardet
from kitchen.text.converters import to_unicode
from openpyxl import Workbook
import xlrd


def write_ws(ws, record, fields):
    """Add a record to a worksheet."""
    new_row = []

    # extract data with field keys from record
    for field in fields:
        new_row.append(record[field])

    # add new row to worksheet
    ws.append(new_row)


def convert_to_float(strValue):
    strValue = strValue.strip().replace(".", "").replace(",", ".")
    floatValue = float(strValue)
    return floatValue


def find_nth(s, x, n):
    i = -1
    for _ in range(n):
        i = s.find(x, i + len(x))
        if i == -1:
            break
    return i


def dict_list_to_excel(dictList, xlName="Output.xlsx"):
    """toma una lista de diccionarios iguales y devuelve una tabla
    en excel con campos tomados de las claves del primer diccionario"""

    # creo el libro y tomo la hoja
    wb = Workbook()
    ws = wb.get_active_sheet()

    # tomo las claves del primer diccionario para usar en todos ellos
    keys = dictList[0].keys()

    # copio los nombres como encabezados
    iCol = 0
    for key in keys:
        ws.cell(row=0, column=iCol).value = key

        iCol += 1

    # itero entre los diccionarios de la lista
    iRow = 1
    for dictionary in dictList:

        iCol = 0
        for key in keys:

            ws.cell(row=iRow, column=iCol).value = dictionary[key]

            iCol += 1

        iRow += 1

    # guarda el excel
    wb.save(xlName)


def get_unicode(string, encoding='utf-8', errors='replace'):
    """fuerza una conversion a unicode a prueba de fallas"""

    # si el valor no es None, intenta convertir a unicode
    if string:
        try:
            RV = to_unicode(string, encoding, errors)

        except Exception:
            encoding = chardet.detect(string)["encoding"]
            RV = to_unicode(string, encoding, errors)

    # si es None, no convierte a unicode
    else:
        RV = string

    return RV


def open_xls_as_xlsx(filename):
    """abre un xls y devuelve un xlsx"""

    # first open using xlrd
    wbXls = xlrd.open_workbook(filename)
    wb = Workbook()

    # borra la hoja inicial del wb creado
    ws = wb.get_active_sheet()
    wb.remove_sheet(ws)

    # agrega cada hoja de un excel al nuevo
    for hoja in wbXls.sheets():

        index = 0
        nrows, ncols = 0, 0
        while nrows * ncols == 0:
            sheet = hoja
            nrows = sheet.nrows
            ncols = sheet.ncols
            index += 1

        # prepare a xlsx sheet
        sheet1 = wb.create_sheet(title = hoja.name)

        for row in xrange(0, nrows):
            for col in xrange(0, ncols):
                sheet1.cell(row=row, column=col).value = sheet.cell_value(row, col)

    return wb
